"""
writer.py
=========
Escritura optimizada para 1,000,000+ pares × 12 semanas.

Salidas:
    resultados_abasto.csv  → formato ANCHO: una fila por Loc-SKU,
                             una columna por fecha (igual que la plantilla)
                             ~15 segundos para 12M valores
    resumen_abasto.xlsx    → Excel pequeño:
                               • Resumen_Clusters
                               • Features_Clusters
                               • Simulacion
                               • Info
"""

import time
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils  import get_column_letter
import pandas as pd
import numpy as np
from typing import List, Tuple


NAVY  = "1B3A6B";  TEAL  = "0D7A7A"
WHITE = "FFFFFF";  LGRAY = "F4F6F8";  MGRAY = "D0D7E3"
LGREEN= "E6F4EC";  RED_L = "FDECEA"

CLUSTER_COLORS = {
    "A": ("E3F2FD", "1565C0"),
    "B": ("FFF3E0", "E65100"),
    "C": ("FCE4EC", "880E4F"),
    "D": ("E8F5E9", "1B5E20"),
}

FEAT_COLS = [
    "Loc","Sku","cluster",
    "media","std","cv","p90","intermitencia","acf1",
    "ratio_demanda_vida","ratio_surtido","margen_relativo","costo_unitario",
    "alpha_cluster","factor_S",
]
FEAT_HEADERS = {
    "Loc":"Loc","Sku":"SKU","cluster":"Cluster",
    "media":"Media (μ)","std":"Std (σ)","cv":"CV",
    "p90":"P90","intermitencia":"Intermitencia","acf1":"ACF1",
    "ratio_demanda_vida":"Ratio D/Vida","ratio_surtido":"Ratio Surtido",
    "margen_relativo":"Margen Relativo","costo_unitario":"Costo Unitario",
    "alpha_cluster":"α cluster","factor_S":"Factor S",
}
FEAT_WIDTHS = [8,8,11,11,11,10,10,14,10,14,14,15,14,11,10]


def _borde():
    s = Side(style="thin", color=MGRAY)
    return Border(left=s, right=s, top=s, bottom=s)

def _cs(ws, row, col, value, bold=False, bg=WHITE, fg="1A1A1A",
        center=False, fmt=None, size=10):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(name="Arial", size=size, bold=bold, color=fg)
    c.fill      = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(
        horizontal="center" if center else "left",
        vertical="center", wrap_text=True)
    c.border = _borde()
    if fmt:
        c.number_format = fmt
    return c

def _header_row(ws, row, headers, widths=None, bg=NAVY, fg=WHITE):
    for ci, h in enumerate(headers, 1):
        _cs(ws, row, ci, h, bold=True, bg=bg, fg=fg, center=True)
    if widths:
        for ci, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[row].height = 28


def _fmt_fecha(v) -> str:
    """Normaliza cualquier tipo de fecha a string dd/mm/yyyy."""
    if isinstance(v, pd.Timestamp):
        return v.strftime("%d/%m/%Y")
    if hasattr(v, 'strftime'):
        return v.strftime("%d/%m/%Y")
    return str(v)


class ResultWriter:
    def __init__(self, verbose=True):
        self.verbose = verbose

    def write(self, resultado_df, features_df,
              fechas_target, sim_stats,
              path_csv, path_xlsx):

        self._write_csv(resultado_df, fechas_target, path_csv)
        self._write_xlsx(features_df, sim_stats, fechas_target, path_xlsx)

    # ── CSV formato ANCHO (igual que la plantilla) ────────────────────────
    def _write_csv(self, resultado_df, fechas_target, path_csv):
        """
        Formato: Loc | Sku | 13/05/2024 | 20/05/2024 | ... | 29/07/2024
        Una fila por par Loc-SKU, una columna por fecha.
        Idéntico a la plantilla de Resultados del Excel.
        """
        t0 = time.time()

        df = resultado_df.copy()

        # Renombrar columnas de fecha a formato dd/mm/yyyy legible
        rename = {}
        for col_orig, ts in fechas_target:
            rename[col_orig] = _fmt_fecha(ts)
        df = df.rename(columns=rename)

        # Asegurar que Loc y Sku sean enteros
        df["Loc"] = df["Loc"].astype(int)
        df["Sku"] = df["Sku"].astype(int)

        # Ordenar por Loc, Sku
        df = df.sort_values(["Loc", "Sku"]).reset_index(drop=True)

        # Escribir CSV con BOM para que Excel lo abra correctamente
        df.to_csv(path_csv, index=False, encoding="utf-8-sig")

        elapsed = time.time() - t0
        n_filas = len(df)
        n_cols  = len(fechas_target)

        if self.verbose:
            print(f"      CSV: {n_filas:,} filas × {n_cols} semanas "
                  f"= {n_filas*n_cols:,} valores")
            print(f"      Guardado: {path_csv}  ({elapsed:.1f}s)")

    # ── Excel resumen (pequeño, con formato) ──────────────────────────────
    def _write_xlsx(self, features_df, sim_stats, fechas_target, path_xlsx):
        t0 = time.time()
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        self._sheet_resumen_clusters(wb, features_df)
        self._sheet_features(wb, features_df)
        self._sheet_simulacion(wb, sim_stats)
        self._sheet_info(wb, features_df, sim_stats, fechas_target)

        wb.save(path_xlsx)
        elapsed = time.time() - t0
        if self.verbose:
            print(f"      Excel resumen: {path_xlsx}  ({elapsed:.1f}s)")

    # ── Hoja 1: Resumen clusters ──────────────────────────────────────────
    def _sheet_resumen_clusters(self, wb, features_df):
        from src.clustering import CLUSTER_PROFILES
        ws = wb.create_sheet("Resumen_Clusters")

        ws.merge_cells("A1:H1")
        c = ws.cell(row=1, column=1,
                    value="RESUMEN DE CLUSTERING — Proyecto Abasto")
        c.font      = Font(name="Arial", size=14, bold=True, color=NAVY)
        c.fill      = PatternFill("solid", fgColor=LGRAY)
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 32

        _header_row(ws, 2,
                    ["Cluster","Perfil","N pares","%",
                     "Media μ","CV","Intermitencia","α cluster"],
                    widths=[10,28,12,8,12,10,14,12], bg=TEAL)

        total = len(features_df)
        for ri, letra in enumerate(sorted(features_df["cluster"].unique()), 3):
            sub    = features_df[features_df["cluster"] == letra]
            n      = len(sub)
            pct    = 100 * n / total
            perfil = CLUSTER_PROFILES.get(letra, {}).get("nombre", "—")
            alpha  = CLUSTER_PROFILES.get(letra, {}).get("alpha_cluster", "—")
            cl_bg, cl_fg = CLUSTER_COLORS.get(letra, (WHITE, "1A1A1A"))

            mu_v  = round(sub["media"].mean(), 2)          if "media"        in sub.columns else "—"
            cv_v  = round(sub["cv"].mean(), 3)              if "cv"           in sub.columns else "—"
            int_v = round(sub["intermitencia"].mean(), 3)   if "intermitencia" in sub.columns else "—"

            vals = [letra, perfil, n, round(pct,1), mu_v, cv_v, int_v, alpha]
            fmts = [None,None,"#,##0","0.0","#,##0.00","0.000","0.000",None]
            for ci, (val, fmt) in enumerate(zip(vals, fmts), 1):
                _cs(ws, ri, ci, val,
                    bold=(ci==1),
                    bg=cl_bg if ci==1 else (LGRAY if ri%2==0 else WHITE),
                    fg=cl_fg if ci==1 else "1A1A1A",
                    center=(ci>=3), fmt=fmt)
            ws.row_dimensions[ri].height = 22

    # ── Hoja 2: Features (sin formato celda-celda para velocidad) ─────────
    def _sheet_features(self, wb, features_df):
        cols = [c for c in FEAT_COLS if c in features_df.columns]
        df   = features_df[cols].copy()
        df.columns = [FEAT_HEADERS.get(c, c) for c in cols]

        ws = wb.create_sheet("Features_Clusters")
        _header_row(ws, 1, list(df.columns),
                    widths=FEAT_WIDTHS[:len(cols)])

        for ri, row in enumerate(df.itertuples(index=False), 2):
            for ci, val in enumerate(row, 1):
                if isinstance(val, np.integer):    val = int(val)
                elif isinstance(val, np.floating): val = round(float(val), 4)
                ws.cell(row=ri, column=ci, value=val)

        ws.freeze_panes = "A2"
        if self.verbose:
            print(f"      Features_Clusters: {len(df):,} filas")

    # ── Hoja 3: Simulación ────────────────────────────────────────────────
    def _sheet_simulacion(self, wb, sim_stats):
        ws = wb.create_sheet("Simulacion")
        ws.merge_cells("A1:C1")
        c = ws.cell(row=1, column=1,
                    value="COMPARATIVO UTILIDAD — Base vs Propuesto")
        c.font      = Font(name="Arial", size=14, bold=True, color=NAVY)
        c.fill      = PatternFill("solid", fgColor=LGRAY)
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 32
        for col, w in zip(["A","B","C"], [32,22,28]):
            ws.column_dimensions[col].width = w
        _header_row(ws, 2, ["Métrica","Valor","Observación"], bg=TEAL)

        ok = sim_stats["delta"] >= 0
        rows = [
            ("Utilidad Esquema Base",
             sim_stats["util_base"],   "$#,##0.00", LGRAY,  "Usa inventario actual I"),
            ("Utilidad Modelo Propuesto",
             sim_stats["util_prop"],   "$#,##0.00", WHITE,  "Usa INV_IDEAL calculado"),
            ("Mejora absoluta ($)",
             sim_stats["delta"],       "$#,##0.00", LGREEN if ok else RED_L, "Propuesto − Base"),
            ("Mejora porcentual (%)",
             sim_stats["mejora_pct"],  "0.00",      LGREEN if ok else RED_L, "% sobre utilidad base"),
            ("Pares × semanas evaluados",
             sim_stats.get("n_eval","—"), "#,##0",  LGRAY,  "Total observaciones"),
        ]
        for ri, (lbl, val, fmt, bg, obs) in enumerate(rows, 3):
            _cs(ws, ri, 1, lbl,  bold=True, bg=bg, size=11)
            _cs(ws, ri, 2, val,  bold=True, bg=bg, center=True, fmt=fmt, size=11)
            _cs(ws, ri, 3, obs,  bg=bg, size=10)
            ws.row_dimensions[ri].height = 24

    # ── Hoja 4: Info ──────────────────────────────────────────────────────
    def _sheet_info(self, wb, features_df, sim_stats, fechas_target):
        from src.clustering import CLUSTER_PROFILES
        ws = wb.create_sheet("Info")
        ws.column_dimensions["A"].width = 28
        ws.column_dimensions["B"].width = 30

        ws.merge_cells("A1:B1")
        c = ws.cell(row=1, column=1, value="INFORMACIÓN DE LA CORRIDA")
        c.font      = Font(name="Arial", size=13, bold=True, color=NAVY)
        c.fill      = PatternFill("solid", fgColor=LGRAY)
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 28

        fechas = [_fmt_fecha(ts) for _, ts in fechas_target]
        info = [
            ("Total pares Loc-SKU",    f"{len(features_df):,}"),
            ("Semanas proyectadas",    len(fechas_target)),
            ("Primera semana",         fechas[0]  if fechas else "—"),
            ("Última semana",          fechas[-1] if fechas else "—"),
            ("Clusters encontrados",   features_df["cluster"].nunique()),
        ]
        for letra in sorted(features_df["cluster"].unique()):
            n   = (features_df["cluster"] == letra).sum()
            nom = CLUSTER_PROFILES.get(letra, {}).get("nombre", "—")
            info.append((f"  Cluster {letra} — {nom}", f"{n:,} pares"))

        for ri, (lbl, val) in enumerate(info, 2):
            _cs(ws, ri, 1, lbl, bold=True,
                bg=LGRAY if ri%2==0 else WHITE)
            _cs(ws, ri, 2, val,
                bg=LGRAY if ri%2==0 else WHITE, center=True)
            ws.row_dimensions[ri].height = 20
